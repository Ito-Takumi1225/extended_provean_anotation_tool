# coding: utf-8
import sys
import xlrd
import re
import subprocess
import openpyxl as px
import os

argv=sys.argv
arg1=argv[1] #path to excelfile
arg2=argv[2] #path to workdirectory
arg3=argv[3] #path to snpEff.jar
arg4=argv[4] #Reference name
arg5=argv[5] #path to provean.sh
arg6=argv[6] #path to scriptdir
excelname=arg1.replace('.xlsx','')  
excelname=re.sub('.*/','',excelname)
print(excelname)

if os.path.exists(arg1):
    pass
else:
    print('python_Error:input file not found')

if os.path.exists(arg2):
    pass
else:
    print('python_Error:workdir not found')

if os.path.exists(arg3):
    pass
else:
    print('python_Error:snpEff.jar not found')
    
if os.path.exists(arg5):
    pass
else:
    print('python_Error:provean not found')

if os.path.exists(arg6):
    pass
else:
    print('scriptdir not found')

f=xlrd.open_workbook(arg1)
sheet=f.sheets()
row1=sheet[0].row_values(0)
row1=list(map(str,row1))

X=row1.index('INFO')
col1=sheet[0].col(X)
col1=list(map(str,col1))
L=len(col1)

book=px.load_workbook(arg1)
sheet1=book.worksheets[0]

py_ver=sys.version
if re.match('3',py_ver):
    py_ver=3
elif re.match('2',py_ver):
    py_ver=2
else:
    pass

if 'PROVEAN_score' in row1:
    Y=row1.index('PROVEAN_score')
else:
    Y=len(row1)
    sheet1.cell(column=Y+1,row=1,value='provean_score')
    sheet1.cell(column=Y+2,row=1,value='provean_pred')

aminodic={'Ala':'A','Arg':'R','Asn':'N','Asp':'D','Cys':'C','Gln':'Q','Glu':'E','Gly':'G','His':'H','Ile':'I','Leu':'L','Lys':'K','Met':'M','Phe':'F','Pro':'P','Ser':'S','Thr':'T','Trp':'W','Tyr':'Y','Val':'V'}
aminokey=aminodic.keys()

for i in range(1,L):
    print(i+1)
    dic1={}
    n0=''
    ls0=[]
    try:
        str1=col1[i]        
        ls1=str1.split('|protein_coding|')
        for j in range(1,len(ls1)):
            ls2=ls1[j].split('|')
            ls3=ls1[j-1].split('|')
            ls2[2]=ls2[2].replace('p.','')
            str2=''
            while str2!=ls2[2]:
                str2=ls2[2]
                for key in aminokey:
                    ls2[2]=ls2[2].replace(key,aminodic[key])
            if ls3[-5]=='HIGH' or ls3[-5]=='MODERATE':
                ls4=[ls3[-1],ls3[-6],ls3[-5],ls2[1],ls2[2]]
                ls0.append(ls4)
            else:
                pass

        if ls0==[]:
            print('pseudogene_or_something')

        else:
            for ls5 in ls0:
                print(ls5)
                n1=''
                try:
                    args=['bash',arg6+'/amino_query_get.sh',ls5[0],arg2,arg3,arg4]
                    res=subprocess.check_output(args)
                except:
                    print('amino query not found')
                
                if py_ver==3:
                    #python3
                    res_python3=str(res)
                    res_python3=res_python3.split('\'')
                    res_python3=res_python3[1]
                    res_python3=res_python3.replace("n",'')
                    res_python3=res_python3.replace("\\",'')
                    res=res_python3
                
                elif py_ver==2:
                    #python2
                    pass

                else:
                    try:
                        #python3
                        res_python3=str(res)
                        res_python3=res_python3.split('\'')
                        res_python3=res_python3[1]
                        res_python3=res_python3.replace("n",'')
                        res_python3=res_python3.replace("\\",'')
                        res=res_python3
                    except:
                        #python2
                        pass
                print(res)

                if ls5[4]=='':
                    str3=ls5[3].replace('c.','')
                    if '*' in str3 or ls5[1]=='sequence_feature':
                        print('sequence feature or etc')

                    else:
                        if re.match('-',str3):
                            m=1
                        else:
                            M=re.search('\d+',str3)
                            m=int(M.group())
                            m=m//3
                        
                        g=open(arg2+'/db/'+excelname+'.var','w')
                        for k in range(m-1,len(res)-1):
                            for amino in aminodic.values():
                                text=res[k]+str(k+1)+amino
                                g.write(text+'\n')
                        g.close()
         
                        try:
                            argsA=['bash', arg6+'/nonpointmutation_score.sh', ls5[0], arg2,  excelname, arg5]
                            resA=subprocess.check_output(argsA)
                            ls6=resA.split()

                            ls7=[]
                            for l in ls6:
                                try:
                                    l=float(l)
                                    ls7.append(l)
                                except:
                                    pass
                            del ls7[0:2]
                            ls7=zip(*[iter(ls7)]*20)

                            ls8=[]
                            for tup in ls7:
                                n1=sum(tup)/20
                                ls8.append(n1)

                            if ls8==[]:
                                pass
                            else:
                                dic1[min(ls8)]='_'+ls5[1]

                        except:
                            print('ERROR:provean_failed')
                
                else:
                    if 'fs' in ls5[4] or '*' in ls5[4]:
                        str4=re.sub('[A-Z]','',ls5[4])
                        str4=str4.replace('fs','')
                        str4=str4.replace('*','')
                        m=int(str4)
                        g=open(arg2+'/db/'+excelname+'.var','w')
                        for k in range(m-1,len(res)-1):
                            for amino in aminodic.values():
                                text=res[k]+str(k+1)+amino
                                g.write(text+'\n')
                        g.close()

                        try:
                            argsA=['bash', arg6+'/nonpointmutation_score.sh', ls5[0], arg2,  excelname, arg5]
                            resA=subprocess.check_output(argsA)
                            ls6=resA.split()

                            ls7=[]
                            for l in ls6:
                                try:
                                    l=float(l)
                                    ls7.append(l)
                                except:
                                    pass
                            del ls7[0:2]
                            ls7=zip(*[iter(ls7)]*20)

                            ls8=[]
                            for tup in ls7:
                                n1=sum(tup)/20
                                ls8.append(n1)
                            
                            if ls8==[]:
                                pass
                            else:
                                dic1[min(ls8)]='_'+ls5[1]

                        except:
                            print('ERROR:provean_failed')

                    else:
                        try:
                            argsB=['bash', arg6+'/pointmutation_score.sh', ls5[0], arg2, excelname, ls5[4], arg5]
                            resB=subprocess.check_output(argsB)
                            dic1[float(resB)]=''
                        except:
                            print('ERROR:provean_failed')

            print(dic1)
    
    except:
        print('ERROR:incorrect cell')

    if dic1=={}:
        pass
    else:
        n0=min(list(dic1.keys()))
        if dic1[n0]=='':
            sheet1.cell(column=Y+1,row=i+1,value=n0)
        else:
            sheet1.cell(column=Y+1,row=i+1,value=str(n0)+dic1[n0])
        
        if n0>-2.5:
            sheet1.cell(column=Y+2,row=i+1,value='N'+dic1[n0])
        else:
            sheet1.cell(column=Y+2,row=i+1,value='D'+dic1[n0])
    book.save(arg2+'/output/out_'+excelname+'.xlsx')  

