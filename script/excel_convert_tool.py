import sys
import openpyxl as px
import subprocess
import os

argv=sys.argv
arg1=argv[1] #path to tmp vcf file
arg2=argv[2] #vcf file name
arg3=argv[3] #path to workdir
vcf_name=arg2.replace('.vcf','')
vcf_name_basic=vcf_name

book=px.Workbook()
sheet1=book.active
sheet1.title='sheet1'

f=open(arg1)
i=1
k=2
for line in f:
    if i==1:
        header=line
        header=header.split('\t')
    else:
        pass
    ls=line.split('\t')
    if i%100000==0:
        book.save(arg3+'/db/'+vcf_name+'.xlsx')
        book=px.Workbook()
        sheet1=book.active
        sheet1.title='sheet1'
        vcf_name=vcf_name_basic+str(k)
        l=1
        for header_str in header:
            sheet1.cell(row=1,column=l,value=header_str)
            l+=1
        k+=1
        i=2
    else:
        pass
    j=1
    for str1 in ls:
        try:
            n=int(str1)
        except:
            try:
                n=float(str1)
            except:
                n=str1
        sheet1.cell(row=i,column=j,value=n)
        j+=1
    i+=1
    
f.close()

book.save(arg3+'/db/'+vcf_name+'.xlsx')
